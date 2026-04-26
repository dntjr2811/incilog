from fastapi import FastAPI, HTTPException, Depends, Query, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy import (
    create_engine, Column, Integer, String, Text, DateTime, Date, Time,
    ForeignKey, text, or_, and_, func
)
from sqlalchemy.orm import sessionmaker, Session, relationship, declarative_base
from pgvector.sqlalchemy import Vector
from pydantic import BaseModel
from datetime import datetime, date, time as time_type
from typing import Optional, List
import os, io, re, logging

import embedding

logger = logging.getLogger(__name__)

# ============================================
# DB設定
# ============================================
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://incilog:incilog_pass@db:5432/incilog")
engine = create_engine(DATABASE_URL)
SessionLocal = sessionmaker(bind=engine)
Base = declarative_base()

# ============================================
# 類似ログ検索の設定
# ============================================
# ============================================
# 類似ログ検索の設定
# ============================================
# cosine類似度の最低ライン（これ未満は除外）
# 環境変数で運用中に調整可能。デフォルト0.75はe5-large日本語ログ用の経験値。
# hostnameを入力から外したことで、意味の解像度が向上し、低めの閾値でも
# ノイズが少なくなる想定。
SIMILARITY_THRESHOLD = float(os.getenv("SIMILARITY_THRESHOLD", "0.75"))

# ============================================
# ORMモデル
# ============================================
class Host(Base):
    __tablename__ = "hosts"
    id = Column(Integer, primary_key=True)
    hostname = Column(String(50), nullable=False, unique=True)
    description = Column(String(200))
    category = Column(String(20), default="other")
    notes = Column(Text)
    created_at = Column(DateTime, default=datetime.utcnow)

class Log(Base):
    __tablename__ = "logs"
    id = Column(Integer, primary_key=True)
    event_date = Column(Date, nullable=False)
    event_time = Column(Time, nullable=False)
    host_id = Column(Integer, ForeignKey("hosts.id"))
    message = Column(Text, nullable=False)
    team = Column(String(20))
    response = Column(Text)
    assignee = Column(String(200))
    reviewer = Column(String(200))
    status = Column(String(20), nullable=False, default="new")
    group_id = Column(Integer)
    message_code = Column(String(30))
    jobnet_path = Column(String(200))
    org_name = Column(String(100))
    source_device = Column(String(50))
    interface_no = Column(String(20))
    event_keyword = Column(String(50))
    jobnet_name = Column(String(200))
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    host = relationship("Host")

class ResponseTemplate(Base):
    __tablename__ = "response_templates"
    id = Column(Integer, primary_key=True)
    title = Column(String(200), nullable=False)
    content = Column(Text, nullable=False)
    team = Column(String(20))
    match_host = Column(String(50))
    match_code = Column(String(30))
    match_device = Column(String(50))
    match_keyword = Column(String(50))
    created_at = Column(DateTime, default=datetime.utcnow)

class LogEmbedding(Base):
    """ログの埋め込みベクトル（log_idをPKとした1:1関連）"""
    __tablename__ = "log_embeddings"
    log_id = Column(Integer, ForeignKey("logs.id", ondelete="CASCADE"), primary_key=True)
    embedding = Column(Vector(embedding.EMBEDDING_DIM), nullable=False)
    model = Column(String(100), nullable=False)
    source_text = Column(Text, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# ============================================
# Pydanticスキーマ
# ============================================
class LogCreate(BaseModel):
    event_date: date
    event_time: str
    hostname: str
    message: str
    team: Optional[str] = None
    jobnet_name: Optional[str] = None
    org_name: Optional[str] = None

class LogUpdate(BaseModel):
    response: Optional[str] = None
    assignee: Optional[str] = None
    reviewer: Optional[str] = None
    status: Optional[str] = None
    team: Optional[str] = None

class HostCreate(BaseModel):
    hostname: str
    description: Optional[str] = None
    category: Optional[str] = "other"
    notes: Optional[str] = None

class TemplateCreate(BaseModel):
    title: str
    content: str
    team: Optional[str] = None
    match_host: Optional[str] = None
    match_code: Optional[str] = None
    match_device: Optional[str] = None
    match_keyword: Optional[str] = None

# ============================================
# メッセージパーサー（自動分類）
# ============================================
def parse_message(msg: str):
    parsed = {"message_code": None, "jobnet_path": None,
              "source_device": None, "interface_no": None, "event_keyword": None}

    code = re.search(r'(KAV[A-Z]\d{4}-[EWI])', msg)
    if code:
        parsed["message_code"] = code.group(1)

    jobnet = re.search(r'(USSSAP[A-Z]\d{4})', msg)
    if jobnet:
        parsed["jobnet_path"] = jobnet.group(1)

    device = re.search(r'\[([A-Z0-9]+)\]', msg)
    if device:
        parsed["source_device"] = device.group(1)

    iface = re.search(r'インターフェイス(\d+)', msg)
    if iface:
        parsed["interface_no"] = iface.group(1)

    kw_map = {
        "linkDown": ["インターフェイスが停止", "linkDown"],
        "linkUp": ["インターフェイスが動作を開始", "linkUp"],
        "vmwVmPoweredOff": ["パワーオフ", "vmwVmPoweredOff"],
        "vmwVmPoweredOn": ["パワーオン", "vmwVmPoweredOn"],
        "vmwVmHBDetected": ["ハートビート", "vmwVmHBDetected"],
        "Preauthentication": ["Preauthentication failed"],
        "coredump": ["systemd-coredump", "dumped core"],
        "grub-boot-success": ["grub-boot-success"],
        "kernel-hung": ["blocked for more than", "hung_task"],
        "QPTZ-Cancel": ["ＱＰＴＺ　Cancel", "QPTZ"],
    }
    for keyword, patterns in kw_map.items():
        if any(p in msg for p in patterns):
            parsed["event_keyword"] = keyword
            break

    return parsed

# ============================================
# 埋め込み生成（バックグラウンドタスク用）
# ============================================
def generate_embedding_task(log_id: int):
    """ログ登録/更新後にバックグラウンドで埋め込みを生成・保存する。

    エラーが発生してもAPIレスポンスには影響させない。
    生成に失敗したログは次回バックフィルスクリプトで補える設計。
    """
    db = SessionLocal()
    try:
        log = db.query(Log).filter(Log.id == log_id).first()
        if not log:
            logger.warning(f"generate_embedding_task: log {log_id} not found")
            return

        hostname = log.host.hostname if log.host else None
        source_text = embedding.build_source_text(hostname, log.message)
        vector = embedding.encode_passage(source_text)

        existing = db.query(LogEmbedding).filter(LogEmbedding.log_id == log_id).first()
        if existing:
            existing.embedding = vector
            existing.model = embedding.EMBEDDING_MODEL_NAME
            existing.source_text = source_text
            existing.updated_at = datetime.utcnow()
        else:
            emb = LogEmbedding(
                log_id=log_id,
                embedding=vector,
                model=embedding.EMBEDDING_MODEL_NAME,
                source_text=source_text,
            )
            db.add(emb)
        db.commit()
        logger.info(f"Embedding generated for log {log_id}")
    except Exception as e:
        logger.error(f"Embedding generation failed for log {log_id}: {e}")
        db.rollback()
    finally:
        db.close()

# ============================================
# アプリケーション
# ============================================
app = FastAPI(title="InciLog v2", version="2.0.0",
              docs_url="/api/docs", openapi_url="/api/openapi.json")

@app.on_event("startup")
def on_startup():
    """起動時に埋め込みモデルをバックグラウンドでウォームアップ。

    モデルの初回ダウンロード（~600MB）がHuggingFaceの応答状況により
    数分かかることがある。同期的にウォームアップするとアプリ起動が
    ブロックされ、healthcheckがタイムアウトする。
    そのため別スレッドで実行し、API自体は即座に起動可能にする。
    モデル未ロード時の埋め込み呼び出しは get_model() 内のロックで
    自動的に待機するため、整合性は保たれる。
    """
    import threading
    threading.Thread(target=embedding.warmup, daemon=True).start()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@app.get("/api/health")
def health(db: Session = Depends(get_db)):
    try:
        db.execute(text("SELECT 1"))
        return {"status": "ok", "db": "connected"}
    except Exception:
        return {"status": "ok", "db": "disconnected"}

# ============================================
# ログ CRUD
# ============================================
@app.get("/api/logs")
def list_logs(status: Optional[str] = None, team: Optional[str] = None,
              host: Optional[str] = None, limit: int = Query(default=100, le=500),
              offset: int = 0, db: Session = Depends(get_db)):
    query = db.query(Log).join(Host, isouter=True).order_by(Log.event_date.desc(), Log.event_time.desc())
    if status:
        query = query.filter(Log.status == status)
    if team:
        query = query.filter(Log.team == team)
    if host:
        query = query.filter(Host.hostname == host)
    total = query.count()
    logs = query.offset(offset).limit(limit).all()
    return {"total": total, "logs": [_log_to_dict(l) for l in logs]}

@app.post("/api/logs", status_code=201)
def create_log(data: LogCreate, background_tasks: BackgroundTasks,
               db: Session = Depends(get_db)):
    host = db.query(Host).filter(Host.hostname == data.hostname).first()
    if not host:
        host = Host(hostname=data.hostname)
        db.add(host)
        db.commit()
        db.refresh(host)
    parsed = parse_message(data.message)
    try:
        t = datetime.strptime(data.event_time, "%H:%M:%S").time()
    except ValueError:
        t = datetime.strptime(data.event_time, "%H:%M").time()
    log = Log(event_date=data.event_date, event_time=t, host_id=host.id,
              message=data.message, team=data.team, org_name=data.org_name,
              jobnet_name=data.jobnet_name, **parsed)
    db.add(log)
    db.commit()
    db.refresh(log)

    # バックグラウンドで埋め込み生成（APIレスポンスを遅延させない）
    background_tasks.add_task(generate_embedding_task, log.id)

    return {"id": log.id, "status": log.status, "parsed": parsed}

@app.get("/api/logs/{log_id}")
def get_log(log_id: int, db: Session = Depends(get_db)):
    log = db.query(Log).filter(Log.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Log not found")
    return _log_to_dict(log)

@app.patch("/api/logs/{log_id}")
def update_log(log_id: int, data: LogUpdate, db: Session = Depends(get_db)):
    log = db.query(Log).filter(Log.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Log not found")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(log, key, value)
    log.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(log)
    # PATCHでは message/hostname を変更しないため埋め込み再生成は不要
    return {"id": log.id, "status": log.status}

# ============================================
# 類似ログ検索（ベクトル類似度ベース、host_idで絞込）
# ============================================
@app.get("/api/logs/{log_id}/similar")
def find_similar(log_id: int, limit: int = 20, db: Session = Depends(get_db)):
    """類似ログ検索（ベクトル類似度のみ）

    検索フロー:
      1. host_id一致を1次フィルタとして必須適用
      2. ベクトル類似度（cosine）で検索、SIMILARITY_THRESHOLD以上のみ採用
      3. response が登録されているログのみ結果に含める
      4. 類似度降順で返す
    """
    target = db.query(Log).filter(Log.id == log_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="Log not found")

    # ホスト未紐付けログは検索対象外（host_id絞込が必須のため）
    if target.host_id is None:
        return _empty_similar_response(target)

    vector_results = _find_vector_matches(
        target=target, limit=limit, db=db,
    )

    similar_logs = []
    for l, sim in vector_results:
        d = _log_to_dict(l)
        d["match_type"] = "vector"
        d["similarity"] = round(float(sim), 4)
        similar_logs.append(d)

    # SQL側で類似度降順に取得済み。念のため再ソート（同値時は新しい日付優先）
    similar_logs.sort(
        key=lambda d: (
            d["similarity"] or 0.0,
            d.get("event_date") or "",
            d.get("event_time") or "",
        ),
        reverse=True,
    )

    return {
        "target_id": target.id,
        "match_criteria": _match_criteria(target),
        "similar_count": len(similar_logs),
        "similar_logs": similar_logs,
    }


def _match_criteria(target: Log) -> dict:
    return {
        "host_id": target.host_id,
        "hostname": target.host.hostname if target.host else None,
        "message_code": target.message_code,
        "jobnet_path": target.jobnet_path,
        "source_device": target.source_device,
        "interface_no": target.interface_no,
        "event_keyword": target.event_keyword,
    }


def _empty_similar_response(target: Log) -> dict:
    return {
        "target_id": target.id,
        "match_criteria": _match_criteria(target),
        "similar_count": 0,
        "similar_logs": [],
    }


def _find_vector_matches(target: Log, limit: int, db: Session) -> List[tuple]:
    """pgvectorによる類似ログ検索。

    - 1次フィルタ: host_id一致（必須）
    - response登録済みのみ
    - cosine類似度 >= SIMILARITY_THRESHOLD のみ
    - 検索クエリは毎回 query: prefix で再エンコード（e5の設計通り）
    """
    # 検索クエリ用ベクトルを query prefix で都度生成
    # （保存済みは passage prefix なので、そのまま使うとe5の意図と異なる）
    try:
        hostname = target.host.hostname if target.host else None
        source_text = embedding.build_source_text(hostname, target.message)
        vec = embedding.encode_query(source_text)
    except Exception as e:
        logger.error(f"Query embedding failed for log {target.id}: {e}")
        return []

    sql = text("""
        SELECT l.id,
               1 - (e.embedding <=> CAST(:vec AS vector)) AS similarity
        FROM logs l
        JOIN log_embeddings e ON e.log_id = l.id
        WHERE l.id != :target_id
          AND l.host_id = :host_id
          AND l.response IS NOT NULL
          AND l.response != ''
          AND (1 - (e.embedding <=> CAST(:vec AS vector))) >= :threshold
        ORDER BY e.embedding <=> CAST(:vec AS vector)
        LIMIT :limit
    """)

    rows = db.execute(sql, {
        "vec": vec,
        "target_id": target.id,
        "host_id": target.host_id,
        "threshold": SIMILARITY_THRESHOLD,
        "limit": limit,
    }).fetchall()

    if not rows:
        return []

    # ID→similarity マップを作り、Logオブジェクト取得後に類似度降順で返す
    id_to_sim = {row.id: row.similarity for row in rows}
    logs = (db.query(Log).join(Host, isouter=True)
            .filter(Log.id.in_(list(id_to_sim.keys()))).all())
    return sorted(
        [(l, id_to_sim[l.id]) for l in logs],
        key=lambda x: x[1],
        reverse=True,
    )

# ============================================
# ホスト管理
# ============================================
@app.get("/api/hosts")
def list_hosts(db: Session = Depends(get_db)):
    hosts = db.query(Host).order_by(Host.hostname).all()
    return [{"id": h.id, "hostname": h.hostname, "description": h.description,
             "category": h.category, "notes": h.notes,
             "log_count": db.query(Log).filter(Log.host_id == h.id).count()} for h in hosts]

@app.post("/api/hosts", status_code=201)
def create_host(data: HostCreate, db: Session = Depends(get_db)):
    if db.query(Host).filter(Host.hostname == data.hostname).first():
        raise HTTPException(status_code=409, detail="Host already exists")
    host = Host(**data.model_dump())
    db.add(host)
    db.commit()
    db.refresh(host)
    return {"id": host.id, "hostname": host.hostname}

# ============================================
# テンプレート管理
# ============================================
@app.get("/api/templates")
def list_templates(team: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(ResponseTemplate).order_by(ResponseTemplate.title)
    if team:
        query = query.filter(ResponseTemplate.team == team)
    return [{"id": t.id, "title": t.title, "content": t.content, "team": t.team,
             "match_code": t.match_code, "match_device": t.match_device,
             "match_keyword": t.match_keyword} for t in query.all()]

@app.post("/api/templates", status_code=201)
def create_template(data: TemplateCreate, db: Session = Depends(get_db)):
    t = ResponseTemplate(**data.model_dump())
    db.add(t)
    db.commit()
    db.refresh(t)
    return {"id": t.id, "title": t.title}

@app.get("/api/logs/{log_id}/suggest-templates")
def suggest_templates(log_id: int, db: Session = Depends(get_db)):
    log = db.query(Log).filter(Log.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Log not found")
    conds = []
    if log.message_code:
        conds.append(ResponseTemplate.match_code == log.message_code)
    if log.event_keyword:
        conds.append(ResponseTemplate.match_keyword == log.event_keyword)
    if not conds:
        return []
    return [{"id": t.id, "title": t.title, "content": t.content}
            for t in db.query(ResponseTemplate).filter(or_(*conds)).all()]

# ============================================
# 統計
# ============================================
@app.get("/api/stats")
def get_stats(db: Session = Depends(get_db)):
    total = db.query(Log).count()
    new_c = db.query(Log).filter(Log.status == "new").count()
    resp_c = db.query(Log).filter(Log.status == "responded").count()
    by_team = db.query(Log.team, func.count(Log.id)).group_by(Log.team).all()
    by_host = (db.query(Host.hostname, func.count(Log.id)).join(Log, Log.host_id == Host.id)
               .group_by(Host.hostname).order_by(func.count(Log.id).desc()).limit(10).all())
    return {"total": total, "new": new_c, "responded": resp_c,
            "closed": total - new_c - resp_c,
            "by_team": {t: c for t, c in by_team if t},
            "top_hosts": {h: c for h, c in by_host}}

# ============================================
# Excel出力
# ============================================
@app.get("/api/logs/export")
def export_logs(date_from: Optional[str] = None, date_to: Optional[str] = None,
                team: Optional[str] = None, db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    query = db.query(Log).join(Host, isouter=True).order_by(Log.event_date, Log.event_time)
    if date_from:
        query = query.filter(Log.event_date >= date_from)
    if date_to:
        query = query.filter(Log.event_date <= date_to)
    if team:
        query = query.filter(Log.team == team)
    logs = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "アラーム一覧"
    hfont = Font(bold=True, size=10)
    hfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    headers = ["No","日付","時刻","ホスト名","メッセージ","担当T","対応内容","担当者","確認者","ステータス"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont; c.fill = hfill; c.border = border
    for i, log in enumerate(logs, 2):
        vals = [log.id, log.event_date.strftime("%Y/%m/%d") if log.event_date else "",
                log.event_time.strftime("%H:%M:%S") if log.event_time else "",
                log.host.hostname if log.host else "", log.message,
                log.team or "", log.response or "", log.assignee or "",
                log.reviewer or "", log.status]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v); c.border = border
    for c, w in zip("ABCDEFGHIJ", [8,12,10,14,60,10,40,12,12,12]):
        ws.column_dimensions[c].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=alarm_logs_{datetime.now().strftime('%Y%m%d')}.xlsx"})

# ============================================
# ヘルパー
# ============================================
def _log_to_dict(l: Log) -> dict:
    return {
        "id": l.id,
        "event_date": l.event_date.isoformat() if l.event_date else None,
        "event_time": l.event_time.strftime("%H:%M:%S") if l.event_time else None,
        "hostname": l.host.hostname if l.host else None,
        "host_description": l.host.description if l.host else None,
        "message": l.message, "team": l.team, "response": l.response,
        "assignee": l.assignee, "reviewer": l.reviewer, "status": l.status,
        "group_id": l.group_id, "message_code": l.message_code,
        "jobnet_path": l.jobnet_path, "org_name": l.org_name,
        "source_device": l.source_device, "interface_no": l.interface_no,
        "event_keyword": l.event_keyword,
    }
